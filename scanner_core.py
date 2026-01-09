import yfinance as yf
import pandas as pd
from ta.momentum import RSIIndicator
import warnings, logging, sys, os
from contextlib import contextmanager
from datetime import datetime
from openpyxl.styles import PatternFill, Border, Side

warnings.simplefilter("ignore", FutureWarning)
logging.getLogger("yfinance").setLevel(logging.ERROR)

# ================= HELPERS =================
def invalid_symbol(df):
    return df is None or df.empty or "Close" not in df.columns

@contextmanager
def suppress_output():
    with open(os.devnull, "w") as devnull:
        old_stdout, old_stderr = sys.stdout, sys.stderr
        sys.stdout, sys.stderr = devnull, devnull
        try:
            yield
        finally:
            sys.stdout, sys.stderr = old_stdout, old_stderr


def get_safe_excel_path(folder):
    base = os.path.join(folder, "nifty500_scan.xlsx")
    if not os.path.exists(base):
        return base
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    return os.path.join(folder, f"nifty500_scan_{ts}.xlsx")


# ================= MAIN SCAN =================
def run_scan(
    REL_RET_MIN,
    REL_RET_MAX,
    RSI_MIN,
    RSI_MAX,
    MIN_VOLUME,
    progress_cb=None
):
    today_folder = datetime.now().strftime("%d %b%y")
    os.makedirs(today_folder, exist_ok=True)

    symbols_df = pd.read_csv("nifty500.csv")
    symbols = [s.strip() + ".NS" for s in symbols_df["Symbol"].dropna().unique()]
    total = len(symbols)

    with suppress_output():
        nifty = yf.download("^NSEI", period="1y", interval="1d",
                            auto_adjust=True, progress=False)

    if isinstance(nifty.columns, pd.MultiIndex):
        nifty.columns = nifty.columns.get_level_values(0)

    nifty_weekly_ret = nifty["Close"].pct_change(5).dropna().iloc[-1] * 100

    rows_all, rows_with_c0, rows_without_c0, rows_with_c0_pullback = [], [], [], []

    for i, sym in enumerate(symbols, start=1):

        if progress_cb:
            progress_cb(sym.replace(".NS", ""), i, total)

        try:
            with suppress_output():
                df = yf.download(sym, period="1y", interval="1d",
                                 auto_adjust=True, progress=False)

            if isinstance(df.columns, pd.MultiIndex):
                df.columns = df.columns.get_level_values(0)

            if invalid_symbol(df):
                continue

            close, low, volume = df["Close"], df["Low"], df["Volume"]

            ema20 = close.ewm(span=20).mean()
            ema50 = close.ewm(span=50).mean()
            ema200 = close.ewm(span=200).mean()
            rsi = RSIIndicator(close).rsi()

            weekly_ret = close.pct_change(5).iloc[-1] * 100
            rel_ret = round(weekly_ret - nifty_weekly_ret, 2)

            close_v = round(close.iloc[-1], 2)
            ema20_v = round(ema20.iloc[-1], 2)
            ema50_v = round(ema50.iloc[-1], 2)
            ema200_v = round(ema200.iloc[-1], 2)
            vol_v = int(volume.iloc[-1])
            rsi_v = round(rsi.iloc[-1], 2)

            c0 = (
                (ema20_v > ema50_v and ema50_v < ema200_v) or
                (ema20_v < ema50_v and ema50_v > ema200_v) or
                (ema20_v > ema50_v and ema50_v > ema200_v)
            )

            base_ok = all([
                close_v > ema20_v,
                close_v > ema50_v,
                vol_v >= MIN_VOLUME,
                RSI_MIN < rsi_v < RSI_MAX,
                REL_RET_MIN < rel_ret < REL_RET_MAX
            ])

            pullback_ok = (low.iloc[-1] <= ema20_v) or (low.iloc[-1] <= ema50_v)

            # ===== ALL STOCKS =====
            row_all = {
                "Stock": sym.replace(".NS", ""),
                "Close": close_v,
                "EMA20": ema20_v,
                "EMA50": ema50_v,
                "EMA200": ema200_v,
                "VOL > 3L": "YES" if vol_v >= MIN_VOLUME else "NO",
                "RSI": rsi_v,
                "Rel%": rel_ret,
                "EMA Structure OK": "YES" if c0 else "NO",
                "Pullback EMA20/50": "YES" if pullback_ok else "NO",
                "STATUS": "OK" if base_ok and c0 and pullback_ok else "NOT OK"
            }
            rows_all.append(row_all)

            # ===== WITHOUT C0 =====
            if base_ok:
                rows_without_c0.append({
                    "Stock": row_all["Stock"],
                    "Close": close_v,
                    "EMA20": ema20_v,
                    "EMA50": ema50_v,
                    "EMA200": ema200_v,
                    "VOL > 3L": row_all["VOL > 3L"],
                    "RSI": rsi_v,
                    "Rel%": rel_ret
                })

            # ===== WITH C0 =====
            if base_ok and c0:
                rows_with_c0.append({
                    "Stock": row_all["Stock"],
                    "Close": close_v,
                    "EMA20": ema20_v,
                    "EMA50": ema50_v,
                    "EMA200": ema200_v,
                    "VOL > 3L": row_all["VOL > 3L"],
                    "RSI": rsi_v,
                    "Rel%": rel_ret,
                    "EMA Structure OK": row_all["EMA Structure OK"]
                })

            # ===== WITH C0 + PULLBACK =====
            if base_ok and c0 and pullback_ok:
                rows_with_c0_pullback.append({
                    "Stock": row_all["Stock"],
                    "Close": close_v,
                    "EMA20": ema20_v,
                    "EMA50": ema50_v,
                    "EMA200": ema200_v,
                    "VOL > 3L": row_all["VOL > 3L"],
                    "RSI": rsi_v,
                    "Rel%": rel_ret,
                    "EMA Structure OK": row_all["EMA Structure OK"],
                    "Pullback EMA20/50": row_all["Pullback EMA20/50"]
                })

        except Exception:
            continue

    dfs = {
        "ALL_STOCKS": pd.DataFrame(rows_all),
        "WITH_C0": pd.DataFrame(rows_with_c0),
        "WITHOUT_C0": pd.DataFrame(rows_without_c0),
        "WITH_C0_PULLBACK": pd.DataFrame(rows_with_c0_pullback),
    }

    for df in dfs.values():
        if not df.empty:
            df.sort_values("Rel%", ascending=False, inplace=True)
            df.reset_index(drop=True, inplace=True)
            df.insert(0, "S.No.", range(1, len(df) + 1))

    # ================= SAFE EXCEL EXPORT =================
    excel_path = get_safe_excel_path(today_folder)

    try:
        writer = pd.ExcelWriter(excel_path, engine="openpyxl")
    except PermissionError:
        excel_path = get_safe_excel_path(today_folder)
        writer = pd.ExcelWriter(excel_path, engine="openpyxl")

    with writer:
        dfs["ALL_STOCKS"].to_excel(writer, "ALL_STOCKS", index=False)
        dfs["WITH_C0"].to_excel(writer, "WITH_C0", index=False)
        dfs["WITHOUT_C0"].to_excel(writer, "WITHOUT_C0", index=False)
        dfs["WITH_C0_PULLBACK"].to_excel(writer, "WITH_C0_PULLBACK", index=False)

        wb = writer.book
        ws = wb["ALL_STOCKS"]

        green = PatternFill("solid", fgColor="C6EFCE")
        red = PatternFill("solid", fgColor="FFC7CE")
        thin = Side(style="thin")
        border = Border(left=thin, right=thin, top=thin, bottom=thin)

        headers = [c.value for c in ws[1]]
        status_col = headers.index("STATUS") + 1

        for r in range(2, ws.max_row + 1):
            fill = green if ws.cell(r, status_col).value == "OK" else red
            for c in range(1, ws.max_column + 1):
                ws.cell(r, c).fill = fill
                ws.cell(r, c).border = border

    return dfs, excel_path
